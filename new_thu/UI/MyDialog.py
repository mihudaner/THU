import json
import os
import shutil
import sys
import subprocess
import numpy as np
import openpyxl
from os import path as osp
from PySide2.QtGui import QIcon, QFont
import pandas as pd
from PySide2.QtCore import Qt
from PySide2.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QAction, QDoubleSpinBox,
    QComboBox, QTableWidget, QTableWidgetItem, QListWidget, QPushButton, QMenu, QCheckBox, QSizePolicy,
    QFormLayout, QTextEdit, QMessageBox, QTreeWidgetItem, QFileDialog, QSpinBox, QSpacerItem
)
from openpyxl.utils import get_column_letter


def get_folder_names(path):
    try:
        # 列出目录中的所有内容
        dir_contents = os.listdir(path)

        # 过滤出文件夹
        folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

        return folders
    except FileNotFoundError:
        print(f"路径 {path} 不存在")
        return []
    except Exception as e:
        print(f"遍历路径 {path} 时发生错误: {e}")
        return []

class MaterialDialog1(QDialog):
    def __init__(self, item):
        super().__init__()
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        self.note = "create"
        self.absolute_path = item.data(0, Qt.UserRole)
        self.setWindowTitle("新建材料")
        self.item = item
        type = self.item.text(0)
        if item.data(0, Qt.UserRole + 1) == "具体材料":
            self.note = "edit"
            type = self.item.parent().text(0)
            self.absolute_path = item.parent().data(0, Qt.UserRole)
            self.setWindowTitle("编辑材料")
        self.setFixedSize(750, 650)
        # 主布局
        self.layout = QVBoxLayout()
        self.layout.setSpacing(15)
        # 名称和类型
        form_layout = QFormLayout()
        self.name_input = QLineEdit()
        self.type_combo = QLineEdit()

        self.type_combo.setText(type)
        self.type_combo.setReadOnly(True)
        form_layout.addRow(QLabel("名称:"), self.name_input)
        form_layout.addRow(QLabel("类型:"), self.type_combo)
        # 添加到布局中
        self.layout.addLayout(form_layout)

        # 材料成分
        self.composition_table = QTableWidget(2, 10)
        self.composition_table.setVerticalHeaderLabels(["元素", "Wt.%"])
        self.layout.addWidget(QLabel("材料成分"))
        self.layout.addWidget(self.composition_table)
        # 设置行高和最大高度
        for row in range(self.composition_table.rowCount()):
            self.composition_table.setRowHeight(row, 40)  # 设置行高为25像素
        self.composition_table.setMaximumHeight(140)  # 设置最大高度为60像素
        # 设置表格的边框和网格线
        self.composition_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid black;  /* 设置表格外边框 */
                gridline-color: gray;     /* 设置网格线颜色 */
            }
            QTableWidget::item {
                border: 1px solid lightgray;  /* 设置单元格边框 */
            }
        """)

        self.cl_layout = QHBoxLayout()
        # 材料属性列表
        self.property_list = QListWidget()
        self.property_list.setFixedWidth(200)
        self.property_list.addItems([
            "密度", "比热", "热导率", "电阻率", "热膨胀系数", "潜热", "弹性", "塑性"
        ])
        # 属性及其单位
        self.property_units = {
            "密度": "kg/m³",
            "比热": "J/g/K",
            "热导率": "W/m/K",
            "电阻率": "Ω∗m",
            "热膨胀系数": "1/K"
        }
        # 创建一个字典来存储每个属性对应的温度表
        temp_table = QTableWidget(100, 2)
        temp_table.setHorizontalHeaderLabels(["温度(K)", "属性值"])
        # Optional: You can populate the table with initial values if needed
        for row in range(100):
            temp_table.setItem(row, 0, QTableWidgetItem(""))
            temp_table.setItem(row, 1, QTableWidgetItem(""))
        self.temp_table = temp_table
        self.temp_tables = {}
        if self.note == "create":
            self.new_type()

        # 当选择材料属性时，更新对应的温度表
        self.property_list.currentItemChanged.connect(self.update_temp_table)

        # layout.addLayout(btn_layout)

        # 温度和属性值表
        self.layout.addWidget(QLabel("材料属性"))
        add_layout = QVBoxLayout()
        add_button = QPushButton("添加")
        del_button = QPushButton("删除")
        add_button_layout = QHBoxLayout()
        add_button_layout.addWidget(add_button)
        add_button_layout.addWidget(del_button)
        add_layout.addWidget(self.property_list)
        add_layout.addLayout(add_button_layout)
        self.cl_layout.addLayout(add_layout)
        self.cl_layout.addWidget(self.temp_table)
        self.layout.addLayout(self.cl_layout)
        # 确定和取消按钮
        button_layout = QHBoxLayout()
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")

        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)

        self.layout.addLayout(button_layout)
        self.setLayout(self.layout)

        # 连接按钮功能
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        add_button.clicked.connect(self.add)
        del_button.clicked.connect(self.delete)
        if self.note == "edit":
            self.initdata()
    def new_type(self):
        for property_name in self.property_list.findItems("*", Qt.MatchWildcard):
            name = property_name.text()

            if name == "潜热":
                temp_table = QTableWidget(100, 3)
                temp_table.setHorizontalHeaderLabels(["潜热（J/kg）", "固相温度(K)", "液相温度（K）"])
                # Optional: You can populate the table with initial values if needed
                for row in range(100):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 示例潜热
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 固相温度
                    temp_table.setItem(row, 2, QTableWidgetItem(""))  # 液相温度
                self.temp_tables[name] = temp_table

            elif name == "弹性":
                temp_table = QTableWidget(100, 2)
                temp_table.setHorizontalHeaderLabels(["弹性模量（MPa）", "泊松比"])
                for row in range(100):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 弹性模量
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 泊松比
                self.temp_tables[name] = temp_table

            elif name == "塑性":
                temp_table = QTableWidget(100, 2)
                temp_table.setHorizontalHeaderLabels(["屈服应力（MPa）", "塑形应变"])
                for row in range(100):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 屈服应力
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 塑形应变
                self.temp_tables[name] = temp_table
            else:
                unit = self.property_units[name]
                temp_table = QTableWidget(100, 2)
                temp_table.setHorizontalHeaderLabels(["温度 (K)", f"{unit}"])
                for row in range(100):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 温度 (K)
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 其他属性值
                self.temp_tables[name] = temp_table

        # 默认显示第一个属性的温度表
        self.current_property = self.property_list.item(0).text()
        self.temp_table = self.temp_tables[self.current_property]
        # 设置表格的边框和网格线
        self.temp_table.setStyleSheet("""
                                    QTableWidget {
                                        border: 1px solid black;  /* 设置表格外边框 */
                                        gridline-color: gray;     /* 设置网格线颜色 */
                                    }
                                    QTableWidget::item {
                                        border: 1px solid lightgray;  /* 设置单元格边框 */
                                    }
                                """)
    def add(self):
        dia = AddPropertyDialog(self)
        if dia.exec_() == QDialog.Accepted:
            self.list_data = dia.list_data
            name = self.list_data[0]
            t1 = self.list_data[1]
            t2 = self.list_data[2]
            t3 = self.list_data[3]
            temp_table = QTableWidget(100, 3)
            temp_table.setHorizontalHeaderLabels([t1, t2, t3])
            for row in range(100):
                temp_table.setItem(row, 0, QTableWidgetItem(""))
                temp_table.setItem(row, 1, QTableWidgetItem(""))
                temp_table.setItem(row, 2, QTableWidgetItem(""))
            self.temp_tables[name] = temp_table
            self.property_list.addItem(name)

    def delete(self):
        # 获取当前选中的项
        current_item = self.property_list.currentItem()
        # 检查是否有选中项，并删除它
        if current_item is not None:
            row = self.property_list.row(current_item)  # 获取选中项的行号
            self.property_list.takeItem(row)  # 删除选中项
            text = current_item.text()
            if text in self.temp_tables:
                del self.temp_tables[text]  # 从字典中删除该键
                print(f"已删除键 '{text}' 对应的值")
    def accept(self):
        material_name = self.name_input.text()
        if not material_name:
            # 弹出提示框
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText("请填写材料名称")
            msg_box.setWindowTitle("警告")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec_()  # 显示弹窗并等待用户点击
            return
        file_path = f"{self.item.data(0, Qt.UserRole)}/{material_name}.xlsx"
        if self.note == "edit":
            # 获取材料名称
            old_name = self.item.text(0)
            if old_name != material_name:
                path = self.item.parent().data(0, Qt.UserRole)
                names = get_folder_names(path)
                if f'{material_name}.xlsx' in names:
                    msg_box = QMessageBox()
                    msg_box.setIcon(QMessageBox.Warning)
                    msg_box.setText("该材料名称已存在，请重新输入")
                    msg_box.setWindowTitle("警告")
                    msg_box.setStandardButtons(QMessageBox.Ok)
                    msg_box.exec_()  # 显示弹窗并等待用户点击
                    return
                old_path = osp.join(self.item.parent().data(0, Qt.UserRole), f'{old_name}.xlsx')
                os.remove(old_path)
            file_path = f"{self.item.parent().data(0, Qt.UserRole)}/{material_name}.xlsx"
        # 创建一个 DataFrame 来存储名称和类型
        name_type_df = pd.DataFrame({
            '名称': [material_name],
            '类型': [self.type_combo.text()]
        })

        # 保存 composition_table 到 sheet1
        composition_data = []
        for row in range(self.composition_table.rowCount()):
            row_data = []
            for col in range(self.composition_table.columnCount()):
                item = self.composition_table.item(row, col)
                row_data.append(item.text() if item else "")
            composition_data.append(row_data)

        # 转置数据
        composition_data = np.array(composition_data).T  # 使用 NumPy 进行转置

        # 创建 DataFrame
        composition_df = pd.DataFrame(composition_data, columns=["元素", "Wt.%"])

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 保存名称和类型到 Excel
            name_type_df.to_excel(writer, sheet_name='名称-类型-成分', index=False, header=False)

            # 将 composition_df 的数据写入，从第三行开始
            composition_df.to_excel(writer, sheet_name='名称-类型-成分', index=False, startrow=1)

            # 获取工作表对象
            worksheet = writer.sheets['名称-类型-成分']
            fixed_width = 20  # 设定的列宽值
            # 遍历所有列并设置列宽
            for column in worksheet.columns:
                column_letter = get_column_letter(column[0].column)  # 获取列字母
                worksheet.column_dimensions[column_letter].width = fixed_width

            # 保存 temp_table 到 sheets
            for property_name in self.property_list.findItems("*", Qt.MatchWildcard):
                name = property_name.text()
                temp_data = []
                temp_table = self.temp_tables[name]
                # 获取所有表头
                header_labels = []
                for column in range(temp_table.columnCount()):
                    header_labels.append(temp_table.horizontalHeaderItem(column).text())
                for row in range(temp_table.rowCount()):
                    row_data = []
                    for col in range(temp_table.columnCount()):
                        # if name in ["潜热", "弹性", "塑性"]:
                        item = temp_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    temp_data.append(row_data)

                # # 为不同属性创建 DataFrame
                temp_df = pd.DataFrame(temp_data, columns=header_labels)
                # temp_df.to_excel(writer, sheet_name=name, index=False)
                try:
                    temp_df.to_excel(writer, sheet_name=name, index=False)
                except Exception as e:
                    print(f"Error saving {name} to {file_path}: {e}")


        print(f"数据已保存到 {file_path}")
        # 新建item
        if self.note == 'edit':
            self.item.setText(0, material_name)
            self.item.setData(0, Qt.UserRole, file_path)
        else:
            new_item = QTreeWidgetItem(self.item)  # 只传入父项
            new_item.setText(0, material_name)
            new_item.setToolTip(0, osp.join(self.absolute_path, f'{material_name}.xlsx'))
            print(f'新建节点absolute_path：{self.absolute_path}')
            new_item.setData(0, Qt.UserRole, osp.join(self.absolute_path, f'{material_name}.xlsx'))
            new_item.setData(0, Qt.UserRole + 1, "具体材料")
            icon_path = osp.join('..', 'resource', 'icon', 'doc.png')
            new_item.setIcon(0, QIcon(icon_path))
        super().accept()  # 调用父类的 accept 方法以关闭对话框

    def initdata(self):
        self.temp_tables = {}
        self.property_list.clear()# 清空 QListWidget 中的所有项目
        # 加载 Excel 文件
        self.workbook = openpyxl.load_workbook(self.item.data(0, Qt.UserRole))
        # 选择工作表
        sheet1 = self.workbook['名称-类型-成分']
        name = self.item.text(0)
        self.name_input.setText(name)
        # 获取所有表名
        sheet_names = [name for name in self.workbook.sheetnames if name != '名称-类型-成分']

        # 读取材料类型和成分并填入self.composition_table
        for row in range(3, 13):  # 从 A5 到 A14
            A = sheet1.cell(row=row, column=1).value  # 读取 A 列
            B = sheet1.cell(row=row, column=2).value  # 读取 B 列
            A_value = str(A) if A is not None else ''
            B_value = str(B) if B is not None else ''
            # 将数据填充到 QTableWidget
            self.composition_table.setItem(0, row - 3, QTableWidgetItem(A_value))  # 设置元素
            self.composition_table.setItem(1, row - 3, QTableWidgetItem(B_value))  # 设置 Wt.%
        # 为每个工作表创建一个 QTableWidget
        flag = 0
        for sheet_name in sheet_names:
            sheet = self.workbook[sheet_name]

            # 根据列数创建 QTableWidget
            column_count = sheet.max_column  # 获取列数

            temp_table = QTableWidget(100, column_count)
            temp_table.setHorizontalHeaderLabels(
                [sheet.cell(row=1, column=i).value for i in range(1, column_count + 1)])  # 表头

            # 填充数据
            for row in range(2, 101):  # 从第二行开始填充数据
                for col in range(1, column_count + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is None:
                        cell_value = ""  # 这里使用空字符串替代 None
                    temp_table.setItem(row - 2, col - 1, QTableWidgetItem(str(cell_value)))  # 填充数据
            self.temp_tables[sheet_name] = temp_table
            self.property_list.addItem(sheet_name)
            if flag == 0:
                flag = 1
                self.current_property = sheet_name
                # 更新显示的 temp_table
                self.cl_layout.itemAt(1).widget().setParent(None)  # 移除旧的temp_table
                self.temp_table = self.temp_tables[self.current_property]
                # 设置表格的边框和网格线
                self.temp_table.setStyleSheet("""
                                                QTableWidget {
                                                    border: 1px solid black;  /* 设置表格外边框 */
                                                    gridline-color: gray;     /* 设置网格线颜色 */
                                                }
                                                QTableWidget::item {
                                                    border: 1px solid lightgray;  /* 设置单元格边框 */
                                                }
                                            """)
                self.cl_layout.insertWidget(1, self.temp_table)  # 在相同位置添加新的temp_table

    def update_temp_table(self, current_item):
        if current_item:
            self.current_property = current_item.text()
            # 更新显示的 temp_table
            self.cl_layout.itemAt(1).widget().setParent(None)  # 移除旧的temp_table
            self.temp_table = self.temp_tables[self.current_property]
            # 设置表格的边框和网格线
            self.temp_table.setStyleSheet("""
                                QTableWidget {
                                    border: 1px solid black;  /* 设置表格外边框 */
                                    gridline-color: gray;     /* 设置网格线颜色 */
                                }
                                QTableWidget::item {
                                    border: 1px solid lightgray;  /* 设置单元格边框 */
                                }
                            """)
            self.cl_layout.insertWidget(1, self.temp_table)  # 在相同位置添加新的temp_table
class MaterialDialog(QDialog):
    def __init__(self, item):
        super().__init__()
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        self.note = "create"
        self.absolute_path = item.data(0, Qt.UserRole)
        self.setWindowTitle("新建材料")
        self.item = item
        type = self.item.text(0)
        if item.data(0, Qt.UserRole + 1) == "具体材料":
            self.note = "edit"
            type = self.item.parent().text(0)
            self.absolute_path = item.parent().data(0, Qt.UserRole)
            self.setWindowTitle("编辑材料")
        self.setFixedSize(750, 650)
        # 主布局
        self.layout = QVBoxLayout()
        self.layout.setSpacing(15)
        # 名称和类型
        form_layout = QFormLayout()
        self.name_input = QLineEdit()
        self.type_combo = QLineEdit()

        self.type_combo.setText(type)
        self.type_combo.setReadOnly(True)
        form_layout.addRow(QLabel("名称:"), self.name_input)
        form_layout.addRow(QLabel("类型:"), self.type_combo)
        # 添加到布局中
        self.layout.addLayout(form_layout)

        # 材料成分
        self.composition_table = QTableWidget(2, 10)
        self.composition_table.setVerticalHeaderLabels(["元素", "Wt.%"])
        self.layout.addWidget(QLabel("材料成分"))
        self.layout.addWidget(self.composition_table)
        # 设置行高和最大高度
        for row in range(self.composition_table.rowCount()):
            self.composition_table.setRowHeight(row, 40)  # 设置行高为25像素
        self.composition_table.setMaximumHeight(140)  # 设置最大高度为60像素
        # 设置表格的边框和网格线
        self.composition_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid black;  /* 设置表格外边框 */
                gridline-color: gray;     /* 设置网格线颜色 */
            }
            QTableWidget::item {
                border: 1px solid lightgray;  /* 设置单元格边框 */
            }
        """)

        self.cl_layout = QHBoxLayout()
        # 材料属性列表
        self.property_list = QListWidget()
        self.property_list.setFixedWidth(200)
        self.property_list.addItems([
            "密度", "比热", "热导率", "电阻率", "热膨胀系数", "潜热", "弹性", "塑性"
        ])
        # 属性及其单位
        self.property_units = {
            "密度": "kg/m³",
            "比热": "J/g/K",
            "热导率": "W/m/K",
            "电阻率": "Ω∗m",
            "热膨胀系数": "1/K"
        }
        # 创建一个字典来存储每个属性对应的温度表
        self.temp_tables = {}
        for property_name in self.property_list.findItems("*", Qt.MatchWildcard):
            name = property_name.text()

            if name == "潜热":
                temp_table = QTableWidget(10, 3)
                temp_table.setHorizontalHeaderLabels(["潜热（J/kg）", "固相温度(K)", "液相温度（K）"])
                # Optional: You can populate the table with initial values if needed
                for row in range(10):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 示例潜热
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 固相温度
                    temp_table.setItem(row, 2, QTableWidgetItem(""))  # 液相温度
                self.temp_tables[name] = temp_table

            elif name == "弹性":
                temp_table = QTableWidget(10, 2)
                temp_table.setHorizontalHeaderLabels(["弹性模量（MPa）", "泊松比"])
                for row in range(10):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 弹性模量
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 泊松比
                self.temp_tables[name] = temp_table

            elif name == "塑性":
                temp_table = QTableWidget(10, 2)
                temp_table.setHorizontalHeaderLabels(["屈服应力（MPa）", "塑形应变"])
                for row in range(10):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))  # 屈服应力
                    temp_table.setItem(row, 1, QTableWidgetItem(""))  # 塑形应变
                self.temp_tables[name] = temp_table
            else:
                unit = self.property_units[name]
                temp_table = QTableWidget(100, 2)
                temp_table.setHorizontalHeaderLabels(["温度 (K)", f"{unit}"])
                for row in range(100):
                    temp_table.setItem(row, 0, QTableWidgetItem(""))   # 温度 (K)
                    temp_table.setItem(row, 1, QTableWidgetItem(""))   # 其他属性值

                self.temp_tables[name] = temp_table

        # 默认显示第一个属性的温度表
        self.current_property = self.property_list.item(0).text()
        self.temp_table = self.temp_tables[self.current_property]
        # 设置表格的边框和网格线
        self.temp_table.setStyleSheet("""
                            QTableWidget {
                                border: 1px solid black;  /* 设置表格外边框 */
                                gridline-color: gray;     /* 设置网格线颜色 */
                            }
                            QTableWidget::item {
                                border: 1px solid lightgray;  /* 设置单元格边框 */
                            }
                        """)

        # 当选择材料属性时，更新对应的温度表
        self.property_list.currentItemChanged.connect(self.update_temp_table)

        # layout.addLayout(btn_layout)

        # 温度和属性值表
        self.layout.addWidget(QLabel("材料属性"))
        self.cl_layout.addWidget(self.property_list)
        self.cl_layout.addWidget(self.temp_table)
        self.layout.addLayout(self.cl_layout)
        # 确定和取消按钮
        button_layout = QHBoxLayout()
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")

        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)

        self.layout.addLayout(button_layout)
        self.setLayout(self.layout)

        # 连接按钮功能
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        if self.note == "edit":
            self.initdata()
    def accept(self):
        # 获取材料名称
        old_name = self.item.text(0)
        material_name = self.name_input.text()
        if not material_name:
            # 弹出提示框
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText("请填写材料名称")
            msg_box.setWindowTitle("警告")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec_()  # 显示弹窗并等待用户点击
            return

        if old_name != material_name:
            path = self.item.parent().data(0, Qt.UserRole)
            names = get_folder_names(path)
            if f'{material_name}.xlsx' in names:
                msg_box = QMessageBox()
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.setText("该材料名称已存在，请重新输入")
                msg_box.setWindowTitle("警告")
                msg_box.setStandardButtons(QMessageBox.Ok)
                msg_box.exec_()  # 显示弹窗并等待用户点击
                return

        file_path = f"{self.absolute_path}/{material_name}.xlsx"
        # 创建一个 DataFrame 来存储名称和类型
        name_type_df = pd.DataFrame({
            '名称': [material_name],
            '类型': [self.type_combo.text()]
        })

        # 保存 composition_table 到 sheet1
        composition_data = []
        for row in range(self.composition_table.rowCount()):
            row_data = []
            for col in range(self.composition_table.columnCount()):
                item = self.composition_table.item(row, col)
                row_data.append(item.text() if item else "")
            composition_data.append(row_data)

        # 转置数据
        composition_data = np.array(composition_data).T  # 使用 NumPy 进行转置

        # 创建 DataFrame
        composition_df = pd.DataFrame(composition_data, columns=["元素", "Wt.%"])

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 保存名称和类型到 Excel
            name_type_df.to_excel(writer, sheet_name='名称-类型-成分', index=False, header=False)

            # 将 composition_df 的数据写入，从第三行开始
            composition_df.to_excel(writer, sheet_name='名称-类型-成分', index=False, startrow=1)

            # 获取工作表对象
            worksheet = writer.sheets['名称-类型-成分']
            fixed_width = 20  # 设定的列宽值
            # 遍历所有列并设置列宽
            for column in worksheet.columns:
                column_letter = get_column_letter(column[0].column)  # 获取列字母
                worksheet.column_dimensions[column_letter].width = fixed_width

            # 保存 temp_table 到 sheets
            for property_name in self.property_list.findItems("*", Qt.MatchWildcard):
                name = property_name.text()
                temp_data = []
                temp_table = self.temp_tables[name]

                for row in range(temp_table.rowCount()):
                    row_data = []
                    for col in range(temp_table.columnCount()):
                        # if name in ["潜热", "弹性", "塑性"]:
                        item = temp_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    temp_data.append(row_data)

                # 为不同属性创建 DataFrame
                if name == "潜热":
                    temp_df = pd.DataFrame(temp_data, columns=["潜热（J/kg）", "固相温度(K)", "液相温度（K）"])
                elif name == "弹性":
                    temp_df = pd.DataFrame(temp_data, columns=["弹性模量（MPa）", "泊松比"])
                elif name == "塑性":
                    temp_df = pd.DataFrame(temp_data, columns=["屈服应力（MPa）", "塑形应变"])
                else:
                    unit = self.property_units[name]
                    temp_df = pd.DataFrame(temp_data, columns=["温度 (K)", f"{unit}"])

                temp_df.to_excel(writer, sheet_name=name, index=False)

        print(f"数据已保存到 {file_path}")
        # message = f"数据已保存到 {file_path}"
        # # 创建消息对话框
        # msg_box = QMessageBox()
        # msg_box.setIcon(QMessageBox.Information)
        # msg_box.setWindowTitle("提示")
        # msg_box.setText(message)
        # msg_box.setStandardButtons(QMessageBox.Ok)
        # # 显示消息对话框
        # msg_box.exec_()
        # 新建item
        if self.note == 'edit':
            self.item.setText(0, material_name)
            self.item.setData(0, Qt.UserRole, file_path)
            old_path = osp.join(self.item.parent().data(0, Qt.UserRole), f'{old_name}.xlsx')
            os.remove(old_path)
            pass
        else:
            new_item = QTreeWidgetItem(self.item)  # 只传入父项
            new_item.setText(0, material_name)
            new_item.setToolTip(0, osp.join(self.absolute_path, f'{material_name}.xlsx'))
            print(f'新建节点absolute_path：{self.absolute_path}')
            new_item.setData(0, Qt.UserRole, osp.join(self.absolute_path, f'{material_name}.xlsx'))
            new_item.setData(0, Qt.UserRole + 1, "具体材料")
            icon_path = osp.join('..', 'resource', 'icon', 'doc.png')
            new_item.setIcon(0, QIcon(icon_path))
        super().accept()  # 调用父类的 accept 方法以关闭对话框

    def initdata(self):

        # 加载 Excel 文件
        self.workbook = openpyxl.load_workbook(self.item.data(0, Qt.UserRole))  # 替换为您的文件路径
        # 选择工作表
        sheet1 = self.workbook['名称-类型-成分']
        name = self.item.text(0)
        self.name_input.setText(name)

        # 读取材料类型和成分并填入self.composition_table
        for row in range(3, 13):  # 从 A5 到 A14
            A = sheet1.cell(row=row, column=1).value  # 读取 A 列
            B = sheet1.cell(row=row, column=2).value  # 读取 B 列
            A_value = str(A) if A is not None else ''
            B_value = str(B) if B is not None else ''
            # 将数据填充到 QTableWidget
            self.composition_table.setItem(0, row - 3, QTableWidgetItem(A_value))  # 设置元素
            self.composition_table.setItem(1, row - 3, QTableWidgetItem(B_value))  # 设置 Wt.%

        sheets = {}
        for property_name in self.property_list.findItems("*", Qt.MatchWildcard):
            name = property_name.text()
            sheets[name] = self.workbook[name]  # 正确使用字典

            # if name == "潜热":
            #     self.load_data_to_table(sheets[name], self.temp_tables[name], range(2, 12))
            if name in ["潜热", "弹性", "塑性"]:
                self.load_data_to_table(sheets[name], self.temp_tables[name], range(2, 12))
            else:
                self.load_data_to_table(sheets[name], self.temp_tables[name], range(2, 102))

    def load_data_to_table(self, sheet, temp_table, row_range):  # 添加 self 参数
        for row in row_range:
            A = sheet.cell(row=row, column=1).value
            B = sheet.cell(row=row, column=2).value

            A_value = str(A) if A is not None else ''
            B_value = str(B) if B is not None else ''

            temp_table.setItem(row - 2, 0, QTableWidgetItem(A_value))
            temp_table.setItem(row - 2, 1, QTableWidgetItem(B_value))

            # 对于潜热需要读取第三列
            if temp_table.columnCount() == 3:
                C = sheet.cell(row=row, column=3).value
                C_value = str(C) if C is not None else ''
                temp_table.setItem(row - 2, 2, QTableWidgetItem(C_value))

    def update_temp_table(self, current_item):
        if current_item:
            self.current_property = current_item.text()
            # 更新显示的 temp_table
            self.cl_layout.itemAt(1).widget().setParent(None)  # 移除旧的temp_table
            self.temp_table = self.temp_tables[self.current_property]
            # 设置表格的边框和网格线
            self.temp_table.setStyleSheet("""
                                QTableWidget {
                                    border: 1px solid black;  /* 设置表格外边框 */
                                    gridline-color: gray;     /* 设置网格线颜色 */
                                }
                                QTableWidget::item {
                                    border: 1px solid lightgray;  /* 设置单元格边框 */
                                }
                            """)
            self.cl_layout.insertWidget(1, self.temp_table)  # 在相同位置添加新的temp_table

class EquipmentDialog(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        self.mainWindow = mainWindow
        self.info_dict = mainWindow.info_dict
        self.item = item
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        self.setWindowTitle("新建设备")
        self.note = "create"
        if self.item.data(0, Qt.UserRole + 1) == "具体设备":
            self.note = "edit"
            self.setWindowTitle("编辑设备")

        if self.note == "edit":
            self.path = item.parent().data(0, Qt.UserRole)
        else:
            self.path = item.data(0, Qt.UserRole)
        self.setFixedSize(600, 600)

        # 主布局
        layout = QVBoxLayout()
        layout.setSpacing(10)
        # 名称和类型
        name_layout = QHBoxLayout()
        self.name_input = QLineEdit()
        if self.note == "edit":
            self.name_input.setText(item.text(0))
            type = item.parent().text(0)
        else:
            type = item.text(0)
        self.type_combo = QLineEdit()
        self.type_combo.setText(type)
        self.type_combo.setReadOnly(True)

        name_layout.addWidget(QLabel("名称:"))
        name_layout.addWidget(self.name_input)
        name_layout.addWidget(QLabel("类型:"))
        name_layout.addWidget(self.type_combo)
        layout.addLayout(name_layout)

        im_layout = QHBoxLayout()
        wen_layout = QHBoxLayout()
        # 设备描述
        self.description_input = QTextEdit()
        self.description_input.setPlaceholderText("设备描述")
        layout.addWidget(QLabel("设备描述:"))
        layout.addWidget(self.description_input)

        # 设备照片管理
        self.image_list = QListWidget()
        # layout.addWidget(self.image_list)

        image_button_layout = QVBoxLayout()
        add_image_button = QPushButton("添加")
        remove_image_button = QPushButton("删除")
        image_button_layout.addWidget(add_image_button)
        image_button_layout.addWidget(remove_image_button)
        im_layout.addWidget(self.image_list)
        im_layout.addLayout(image_button_layout)

        # 设备文件管理
        self.file_list = QListWidget()
        # layout.addWidget(self.file_list)

        file_button_layout = QVBoxLayout()
        add_file_button = QPushButton("添加")
        remove_file_button = QPushButton("删除")
        file_button_layout.addWidget(add_file_button)
        file_button_layout.addWidget(remove_file_button)
        wen_layout.addWidget(self.file_list)
        wen_layout.addLayout(file_button_layout)
        layout.addWidget(QLabel("设备照片:"))
        layout.addLayout(im_layout)
        layout.addWidget(QLabel("设备文件:"))
        layout.addLayout(wen_layout)
        # 确定和取消按钮
        button_layout = QHBoxLayout()
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

        # 按钮功能（示例）
        add_image_button.clicked.connect(self.add_image)
        remove_image_button.clicked.connect(self.remove_image)
        add_file_button.clicked.connect(self.add_file)
        remove_file_button.clicked.connect(self.remove_file)
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        if self.note == "edit":
            self.initData()

    def get_folder_names(self, path):
        try:
            # 列出目录中的所有内容
            dir_contents = os.listdir(path)

            # 过滤出文件夹
            folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

            return folders
        except FileNotFoundError:
            print(f"路径 {path} 不存在")
            return []
        except Exception as e:
            print(f"遍历路径 {path} 时发生错误: {e}")
            return []
    def initData(self):
        # filepath = osp.join(self.item.data(0, Qt.UserRole), self.name_input.text())
        filepath = osp.join(self.item.data(0, Qt.UserRole), "设备描述.txt")
        # 确保文件存在
        if not osp.exists(filepath):
            QMessageBox.warning(self, "警告", f"文件 {filepath} 不存在")
            return
        try:
            # 以读模式打开文件并读取内容
            with open(filepath, 'r', encoding='utf-8') as file:
                content = file.read()  # 读取文件内容
            # 将内容设置到 QTextEdit
            self.description_input.setPlainText(content)  # 使用 setPlainText()方法
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法读取文件: {e}")
        tu_dir = osp.join(self.item.data(0, Qt.UserRole), '设备图片')
        # 列出目录下的所有文件
        for filename in os.listdir(tu_dir):
            self.image_list.addItem(filename)  # 添加文件名到 QListWidget
        wen_dir = osp.join(self.item.data(0, Qt.UserRole), '设备文件')
        # 列出目录下的所有文件
        for filename in os.listdir(wen_dir):
            self.file_list.addItem(filename)  # 添加文件名到 QListWidget

    def accept(self):
        # 名称检查
        name = self.name_input.text()
        names = self.get_folder_names(self.path)
        if not self.name_input.text():
            QMessageBox.warning(self, "警告", "请输入设备名称")
            return
        # 新建模式先创建文件
        if self.note == "edit":
            ori_name = self.item.text(0)
            if name != ori_name:
                if name in names:
                    QMessageBox.warning(self, "警告", "该设备已存在")
                    return
                os.rename(osp.join(self.path, ori_name), osp.join(self.path, name))
                self.item.setText(0, name)
                self.item.setData(0, Qt.UserRole, osp.join(self.path, name))
        else:
            if name in names:
                QMessageBox.warning(self, "警告", "该设备名已存在")
                return
            new_item = QTreeWidgetItem(self.item)
            new_item.setText(0, name)
            new_item.setData(0, Qt.UserRole, osp.join(self.path, name))
            new_item.setData(0, Qt.UserRole + 1, "具体设备")
            # 设置窗口图标
            icon_path = os.path.join('..', 'resource', 'icon', 'dir.png')
            if os.path.exists(icon_path):
                new_item.setIcon(0, QIcon(icon_path))
            else:
                print(f"Icon file not found: {icon_path}")
            dir_path = osp.join(self.path, name)
            os.makedirs(dir_path, exist_ok=True)
            # 新建图片和文件dir
            miaoshu = osp.join(dir_path, '设备描述.txt')
            tu_dir = osp.join(dir_path, '设备图片')
            wen_dir = osp.join(dir_path, '设备文件')
            try:
                # 创建设备图片和设备文件的文件夹
                os.makedirs(tu_dir, exist_ok=True)  # exist_ok=True表示如果文件夹已存在，不会抛出异常
                os.makedirs(wen_dir, exist_ok=True)
                # 创建一个空的文本文件
                with open(miaoshu, 'w', encoding='utf-8') as file:
                    file.write('')  # 可以在文件中写入一些初始内容，如果需要的话
            except Exception as e:
                print(f'创建文件夹或文件时出错: {e}')
            for obj in os.listdir(new_item.data(0, Qt.UserRole)):
                tmp_path = osp.join(new_item.data(0, Qt.UserRole), obj)
                if osp.isdir(tmp_path):
                    dir_item = self.mainWindow._generate_item(new_item, obj, tmp_path, 0)
                    self.mainWindow.list_dir(dir_item, tmp_path)
                else:
                    self.mainWindow._generate_item(new_item, obj, tmp_path, 1)

        description = self.description_input.toPlainText()
        # file = osp.join(self.item.data(0, Qt.UserRole), self.name_input.text())
        file = osp.join(self.path, name, "设备描述")
        filepath = f"{file}.txt"  # 添加文件扩展名
        try:
            with open(filepath, 'w', encoding='utf-8') as file:  # 以写入模式打开文件
                file.write(description)  # 将描述写入文件
            # QMessageBox.information(self, '成功', f'描述已保存到 {filepath}')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'无法保存文件: {e}')
        tu_dir = osp.join(self.path, name, '设备图片')
        wen_dir = osp.join(self.path, name, '设备文件')
        # 遍历 QListWidget 中的所有项
        for index in range(self.image_list.count()):
            item = self.image_list.item(index)  # 获取 QListWidget 中的项
            image_filename = item.text()  # 获取文件名
            # 检查文件路径是否为绝对路径
            if not os.path.isabs(image_filename):
                pass
            else:
                try:
                    # 复制文件到目标目录
                    shutil.copy(image_filename, tu_dir)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"无法复制文件 {image_filename}: {e}")
        # 遍历 QListWidget 中的所有项
        for index in range(self.file_list.count()):
            item = self.file_list.item(index)  # 获取 QListWidget 中的项
            image_filename = item.text()  # 获取文件名
            # 检查文件路径是否为绝对路径
            if not os.path.isabs(image_filename):
                pass
            else:
                try:
                    # 复制文件到目标目录
                    shutil.copy(image_filename, wen_dir)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"无法复制文件 {image_filename}: {e}")

        super().accept()

    def add_image(self):
        # 打开文件对话框选择图片文件
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "选择文件", "",
                                                "All Files (*)",  # 只使用 "All Files (*)"
                                                options=options)
        if files:  # 如果用户选择了文件
            for file in files:
                self.image_list.addItem(file)  # 将文件路径添加到 QListWidget

    def remove_image(self):
        # 获取选中的项并删除
        selected_items = self.image_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要删除的项目")
            return
        for item in selected_items:
            self.image_list.takeItem(self.image_list.row(item))  # 从列表中删除选中的项

    def add_file(self):
        # 打开文件对话框选择文件
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "选择文件", "",
                                                "All Files (*)",  # 只使用 "All Files (*)"
                                                options=options)
        if files:  # 如果用户选择了文件
            for file in files:
                self.file_list.addItem(file)  # 将文件路径添加到 QListWidget

    def remove_file(self):
        # 获取选中的项并删除
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择要删除的项目")
            return

        for item in selected_items:
            self.file_list.takeItem(self.file_list.row(item))  # 从列表中删除选中的项

class ProcessDialog(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        self.item = item
        self.note = "create"
        if self.item.data(0, Qt.UserRole + 1) == "具体工艺":
            self.note = "edit"
            self.setWindowTitle(f"{item.parent().text(0)}工艺")
            self.style = item.parent().text(0)
        else:
            self.setWindowTitle(f"{item.text(0)}工艺")
            self.style = item.text(0)
        self.path = item.data(0, Qt.UserRole)
        self.setFixedSize(650, 350)
        # 设置窗口图标（可选）
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        # 创建主布局
        layout = QVBoxLayout()
        form = QHBoxLayout()
        form_layout1 = QFormLayout()
        form_layout1.setSpacing(13)  # 设置表单布局中控件间隙
        form_layout2 = QFormLayout()
        form_layout2.setSpacing(13)  # 设置表单布局中控件间隙

        # 创建输入字段
        self.name = QLineEdit()
        self.material_edit = QComboBox()
        self.material_edit.addItems(mainWindow.info_dict["具体材料"])
        form_layout1.addRow("工艺名称", self.name)
        form_layout1.addRow("熔覆材料", self.material_edit)

        self.base_material_edit = QComboBox()
        self.base_material_edit.addItems(mainWindow.info_dict["具体材料"])
        form_layout2.addRow("   ", QLabel("  "))
        form_layout2.addRow("基板材料", self.base_material_edit)

        self.fmd_edit = QLineEdit()
        # self.fmd_edit.setRange(0, 1000)  # 设置范围 粉末粒径(μm)
        self.bansize_edit = QLineEdit()
        # self.bansize_edit.setRange(0, 500)  # 设置范围 基板尺寸(mm)
        self.scd_edit = QDoubleSpinBox()
        self.scd_edit.setRange(0, 10000)  # 设置范围 丝材直径(mm)
        if self.style == "送粉":
            form_layout1.addRow("粉末粒径(μm)", self.fmd_edit)
            form_layout2.addRow("基板尺寸(mm)", self.bansize_edit)
        else:
            form_layout1.addRow("丝材直径(mm)", self.scd_edit)
            form_layout2.addRow("基板尺寸(mm)", self.bansize_edit)

        self.laser_power_edit = QSpinBox()
        self.laser_power_edit.setRange(0, 10000)  # 设置范围 激光功率(W)
        form_layout1.addRow("激光功率(W)", self.laser_power_edit)
        self.welding_speed_edit = QDoubleSpinBox()
        self.welding_speed_edit.setRange(0, 10000)  # 设置范围 熔覆速度(mm/s)
        form_layout2.addRow("熔覆速度(mm/s)", self.welding_speed_edit)

        self.sf_rate_edit = QDoubleSpinBox()
        self.sf_rate_edit.setRange(0, 10000)  # 设置范围 送粉转速(r/min)
        self.ss_rate_edit = QDoubleSpinBox()
        self.ss_rate_edit.setRange(0, 10000)  # 设置范围 送丝速度(m/min)
        if self.style == "送粉":
            form_layout1.addRow("送粉转速(r/min)", self.sf_rate_edit)
        else:
            form_layout1.addRow("送丝速度(m/min)", self.ss_rate_edit)

        self.addition_rate_edit = QDoubleSpinBox()
        self.addition_rate_edit.setRange(0, 10000)  # 设置范围 质量添加(g/min)
        form_layout2.addRow("质量添加(g/min)", self.addition_rate_edit)

        self.spot_voltage_edit = QDoubleSpinBox()
        self.spot_voltage_edit.setRange(0, 10000)  # 设置范围 光斑电压(V)
        self.spot_diameter_edit = QDoubleSpinBox()
        self.spot_diameter_edit.setRange(0, 10000)  # 设置范围 光斑直径(mm)
        self.gap_interval_edit = QDoubleSpinBox()
        self.gap_interval_edit.setRange(0, 10000)  # 设置范围 道间间隔(s)
        self.layer_interval_edit = QDoubleSpinBox()
        self.layer_interval_edit.setRange(0, 10000)  # 设置范围 层间间隔(s)

        if self.style == "送粉":
            form_layout1.addRow("光斑电压(V)", self.spot_voltage_edit)
            form_layout1.addRow("道间间隔(s)", self.gap_interval_edit)
            form_layout2.addRow("光斑直径(mm)", self.spot_diameter_edit)
            form_layout2.addRow("层间间隔(s)", self.layer_interval_edit)

        self.offset_edit = QDoubleSpinBox()
        self.offset_edit.setRange(0, 10000)  # 设置范围 道间偏移(mm)
        form_layout1.addRow("道间偏移(mm)", self.offset_edit)

        self.lift_height_edit = QDoubleSpinBox()
        self.lift_height_edit.setRange(0, 10000)  # 设置范围 层间抬升(mm)
        form_layout2.addRow("层间抬升(mm)", self.lift_height_edit)

        self.protect_gas_flow_edit = QLineEdit()
        # self.protect_gas_flow_edit.setRange(0, 10000)  # 设置范围 保护气及流量(L/min)

        self.carrier_gas_flow_edit = QLineEdit()
        # self.carrier_gas_flow_edit.setRange(0, 10000)  # 设置范围 载气及流量(L/min)

        self.qd_flow_edit = QLineEdit()
        # self.qd_flow_edit.setRange(0, 10000)  # 设置范围 气刀及流量(L/min)

        self.pre_time_edit = QDoubleSpinBox()
        self.pre_time_edit.setRange(0, 10000)  # 设置范围 加工前保护气时长(s)
        self.keep_time_edit = QDoubleSpinBox()
        self.keep_time_edit.setRange(0, 10000)  # 设置范围 保护气保持时间(s)

        if self.style == "送粉":
            form_layout1.addRow("保护气及流量(L/min)", self.protect_gas_flow_edit)
            form_layout2.addRow("载气及流量(L/min)", self.carrier_gas_flow_edit)
        else:
            form_layout1.addRow("保护气及流量(L/min)", self.protect_gas_flow_edit)
            form_layout2.addRow("气刀及流量(L/min)", self.qd_flow_edit)
            form_layout1.addRow("加工前保护气时长(s)", self.pre_time_edit)
            form_layout2.addRow("保护气保持时间(s)", self.keep_time_edit)

        form.addLayout(form_layout1)
        form.addLayout(form_layout2)
        layout.addLayout(form)
        # 确定和取消按钮
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("确认")
        self.cancel_button = QPushButton("取消")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
        # 设置主布局
        self.setLayout(layout)
        # 连接信号
        self.ok_button.clicked.connect(self.save_to_json)
        self.cancel_button.clicked.connect(self.reject)

        if self.note == "edit":
            self.initData()
    def initData(self):
        name = self.item.text(0)
        self.name.setText(name)
        if os.path.exists(self.path):
            with open(self.path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # 填入控件
            if self.style == "送粉":
                self.material_edit.setCurrentText(data.get("熔覆材料", ""))
                self.base_material_edit.setCurrentText(data.get("基板材料", ""))
                self.fmd_edit.setText(data.get("粉末粒径(μm)", ""))
                self.bansize_edit.setText(data.get("基板尺寸(mm)", ""))
                self.laser_power_edit.setValue(data.get("激光功率(W)", 0))
                self.welding_speed_edit.setValue(data.get("熔覆速度(mm/s)", 0))
                self.sf_rate_edit.setValue(data.get("送粉转速(r/min)", 0))
                self.addition_rate_edit.setValue(data.get("质量添加(g/min)", 0))
                self.spot_voltage_edit.setValue(data.get("光斑电压(V)", 0))
                self.spot_diameter_edit.setValue(data.get("光斑直径(mm)", 0))
                self.gap_interval_edit.setValue(data.get("道间间隔(s)", 0))
                self.layer_interval_edit.setValue(data.get("层间间隔(s)", 0))
                self.offset_edit.setValue(data.get("道间偏移(mm)", 0))
                self.lift_height_edit.setValue(data.get("层间抬升(mm)", 0))
                self.protect_gas_flow_edit.setText(data.get("保护气及流量(L/min)", ""))
                self.carrier_gas_flow_edit.setText(data.get("载气及流量(L/min)", ""))
            else:
                self.material_edit.setCurrentText(data.get("熔覆材料", ""))
                self.base_material_edit.setCurrentText(data.get("基板材料", ""))
                self.scd_edit.setValue(data.get("丝材直径(mm)", 0))
                self.bansize_edit.setText(data.get("基板尺寸(mm)", ""))
                self.laser_power_edit.setValue(data.get("激光功率(W)", 0))
                self.welding_speed_edit.setValue(data.get("熔覆速度(mm/s)", 0))
                self.ss_rate_edit.setValue(data.get("送丝速度(m/min)", 0))
                self.addition_rate_edit.setValue(data.get("质量添加(g/min)", 0))
                self.offset_edit.setValue(data.get("道间偏移(mm)", 0))
                self.lift_height_edit.setValue(data.get("层间抬升(mm)", 0))
                self.protect_gas_flow_edit.setText(data.get("保护气及流量(L/min)", ""))
                self.qd_flow_edit.setText(data.get("气刀及流量(L/min)", ""))
                self.pre_time_edit.setValue(data.get("加工前保护气时长(s)", 0))
                self.keep_time_edit.setValue(data.get("保护气保持时间(s)", 0))

    def get_folder_names(self, path):
        try:
            # 列出目录中的所有内容
            dir_contents = os.listdir(path)

            # 过滤出文件夹
            folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

            return folders
        except FileNotFoundError:
            print(f"路径 {path} 不存在")
            return []
        except Exception as e:
            print(f"遍历路径 {path} 时发生错误: {e}")
            return []
    def save_to_json(self):
        # 名称检查
        if self.note == "edit":
            path = self.item.parent().data(0, Qt.UserRole)
        else:
            path = self.item.data(0, Qt.UserRole)
        names = self.get_folder_names(path)
        if not self.name.text():
            QMessageBox.warning(self, "警告", "请输入工艺名称")
            return

        if self.note == "edit":
            ori_name = self.item.text(0)
            if self.name.text() != ori_name:
                if self.name.text() in names:
                    QMessageBox.warning(self, "警告", "该工艺名已存在")
                    return
            if os.path.exists(self.path):
                os.remove(self.path)
            file_path = osp.join(self.item.parent().data(0, Qt.UserRole), f"{self.name.text()}.json")
            self.item.setText(0, self.name.text())
            self.item.setData(0, Qt.UserRole, file_path)
            type = self.item.parent().text(0)
        else:
            if self.name.text() in names:
                QMessageBox.warning(self, "警告", "该工艺名已存在")
                return
            new_item = QTreeWidgetItem(self.item)
            file_path = osp.join(self.item.data(0, Qt.UserRole), f"{self.name.text()}.json")
            new_item.setText(0, self.name.text())
            new_item.setData(0, Qt.UserRole, file_path)
            new_item.setData(0, Qt.UserRole + 1, "具体工艺")
            # 设置图标
            icon_path = os.path.join('..', 'resource', 'icon', 'doc.png')
            new_item.setIcon(0, QIcon(icon_path))
        if self.style == "送粉":
            data = {
                "工艺名称": self.name.text(),
                "熔覆材料": self.material_edit.currentText(),
                "基板材料": self.base_material_edit.currentText(),
                "粉末粒径(μm)": self.fmd_edit.text(),
                "基板尺寸(mm)": self.bansize_edit.text(),
                "激光功率(W)": self.laser_power_edit.value(),
                "熔覆速度(mm/s)": self.welding_speed_edit.value(),
                "送粉转速(r/min)": self.ss_rate_edit.value(),
                "质量添加(g/min)": self.addition_rate_edit.value(),
                "光斑电压(V)": self.spot_voltage_edit.value(),
                "光斑直径(mm)": self.spot_diameter_edit.value(),
                "道间间隔(s)": self.gap_interval_edit.value(),
                "层间间隔(s)": self.layer_interval_edit.value(),
                "道间偏移(mm)": self.offset_edit.value(),
                "层间抬升(mm)": self.lift_height_edit.value(),
                "保护气及流量(L/min)": self.protect_gas_flow_edit.text(),
                "载气及流量(L/min)": self.carrier_gas_flow_edit.text()
            }
        else:
            data = {
                "工艺名称": self.name.text(),
                "熔覆材料": self.material_edit.currentText(),
                "基板材料": self.base_material_edit.currentText(),
                "丝材直径(mm)": self.scd_edit.value(),
                "基板尺寸(mm)": self.bansize_edit.text(),
                "激光功率(W)": self.laser_power_edit.value(),
                "熔覆速度(mm/s)": self.welding_speed_edit.value(),
                "送丝速度(m/min)": self.ss_rate_edit.value(),
                "质量添加(g/min)": self.addition_rate_edit.value(),
                "道间偏移(mm)": self.offset_edit.value(),
                "层间抬升(mm)": self.lift_height_edit.value(),
                "保护气及流量(L/min)": self.protect_gas_flow_edit.text(),
                "气刀及流量(L/min)": self.qd_flow_edit.text(),
                "加工前保护气时长(s)": self.pre_time_edit.value(),
                "保护气保持时间(s)": self.keep_time_edit.value()
            }
        if file_path:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            self.accept()

class ProcessDialog2(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        self.item = item
        self.note = "create"
        if self.item.data(0, Qt.UserRole + 1) == "具体工艺":
            self.note = "edit"
        self.path = item.data(0, Qt.UserRole)
        self.setWindowTitle("粉丝同送")
        self.setFixedSize(670, 380)
        # 设置窗口图标（可选）
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        # 创建主布局
        layout = QVBoxLayout()
        form = QHBoxLayout()
        form_layout1 = QFormLayout()
        form_layout1.setSpacing(13)  # 设置表单布局中控件间隙
        form_layout2 = QFormLayout()
        form_layout2.setSpacing(13)  # 设置表单布局中控件间隙

        # 创建输入字段
        self.name = QLineEdit()  # 工艺名称
        self.material_edit = QComboBox()  # 粉末材料
        self.material_edit.addItems(mainWindow.info_dict["具体材料"])
        self.material_edit2 = QComboBox()  # 送丝材料
        self.material_edit2.addItems(mainWindow.info_dict["具体材料"])

        self.base_material_edit = QComboBox()  # 基板材料
        self.base_material_edit.addItems(mainWindow.info_dict["具体材料"])

        self.sf_rate_edit = QDoubleSpinBox()
        self.sf_rate_edit.setRange(0, 10000)  # 设置范围 送粉转速(r/min)
        self.ss_rate_edit = QDoubleSpinBox()
        self.ss_rate_edit.setRange(0, 10000)  # 设置范围 送丝速度(m/min)
        self.fmd_edit = QLineEdit()  # 粉末粒径
        self.bansize_edit = QLineEdit()  # 基板尺寸
        self.scd_edit = QDoubleSpinBox()  # 丝材直径
        self.scd_edit.setRange(0, 10000)  # 设置范围 丝材直径(mm)
        self.addition_rate_edit = QDoubleSpinBox()
        self.addition_rate_edit.setRange(0, 10000)  # 设置范围 质量添加(g/min)
        self.addition_rate_edit2 = QDoubleSpinBox()
        self.addition_rate_edit2.setRange(0, 10000)  # 设置范围 质量添加(g/min)

        self.spot_voltage_edit = QDoubleSpinBox()
        self.spot_voltage_edit.setRange(0, 10000)  # 设置范围 光斑电压(V)
        self.spot_diameter_edit = QDoubleSpinBox()
        self.spot_diameter_edit.setRange(0, 10000)  # 设置范围 光斑直径(mm)
        self.gap_interval_edit = QDoubleSpinBox()
        self.gap_interval_edit.setRange(0, 10000)  # 设置范围 道间间隔(s)
        self.layer_interval_edit = QDoubleSpinBox()
        self.layer_interval_edit.setRange(0, 10000)  # 设置范围 层间间隔(s)

        self.laser_power_edit = QSpinBox()
        self.laser_power_edit.setRange(0, 10000)  # 设置范围 激光功率(W)
        self.welding_speed_edit = QDoubleSpinBox()
        self.welding_speed_edit.setRange(0, 10000)  # 设置范围 熔覆速度(mm/s)

        self.protect_gas_flow_edit = QLineEdit()
        # self.protect_gas_flow_edit.setRange(0, 10000)  # 设置范围 保护气及流量(L/min)
        self.carrier_gas_flow_edit = QLineEdit()
        # self.carrier_gas_flow_edit.setRange(0, 10000)  # 设置范围 载气及流量(L/min)
        self.qd_flow_edit = QLineEdit()
        # self.qd_flow_edit.setRange(0, 10000)  # 设置范围 气刀及流量(L/min)
        self.pre_time_edit = QDoubleSpinBox()
        self.pre_time_edit.setRange(0, 10000)  # 设置范围 加工前保护气时长(s)
        self.keep_time_edit = QDoubleSpinBox()
        self.keep_time_edit.setRange(0, 10000)  # 设置范围 保护气保持时间(s)
        self.offset_edit = QDoubleSpinBox()
        self.offset_edit.setRange(0, 10000)  # 设置范围 道间偏移(mm)
        self.lift_height_edit = QDoubleSpinBox()
        self.lift_height_edit.setRange(0, 10000)  # 设置范围 层间抬升(mm)
        form_layout1.addRow("工艺名称", self.name)
        form_layout1.addRow("粉末材料", self.material_edit)
        form_layout1.addRow("粉末粒径(μm)", self.fmd_edit)
        form_layout1.addRow("送粉转速(r/min)", self.sf_rate_edit)
        form_layout1.addRow("粉末质量添加(g/min)", self.addition_rate_edit)
        form_layout1.addRow("粉末载气及流量(L/min)", self.carrier_gas_flow_edit)
        form_layout1.addRow("送丝材料", self.material_edit2)
        form_layout1.addRow("丝材直径(mm)", self.scd_edit)
        form_layout1.addRow("送丝速度(m/min)", self.ss_rate_edit)
        form_layout1.addRow("丝质量添加(g/min)", self.addition_rate_edit2)

        form_layout2.addRow("基板材料", self.base_material_edit)
        form_layout2.addRow("基板尺寸(mm)", self.bansize_edit)
        form_layout2.addRow("激光功率(W)", self.laser_power_edit)
        form_layout2.addRow("熔覆速度(mm/s)", self.welding_speed_edit)
        form_layout2.addRow("道间偏移(mm)", self.offset_edit)
        form_layout2.addRow("层间抬升(mm)", self.lift_height_edit)
        form_layout2.addRow("保护气及流量(L/min)", self.protect_gas_flow_edit)
        form_layout2.addRow("保护气保持时间(s)", self.keep_time_edit)
        form_layout2.addRow("气刀及流量(L/min)", self.qd_flow_edit)
        form_layout2.addRow("加工前保护气时长(s)", self.pre_time_edit)

        form.addLayout(form_layout1)
        form.addLayout(form_layout2)
        layout.addLayout(form)
        # 确定和取消按钮
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("确认")
        self.cancel_button = QPushButton("取消")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
        # 设置主布局
        self.setLayout(layout)
        # 连接信号
        self.ok_button.clicked.connect(self.save_to_json)
        self.cancel_button.clicked.connect(self.reject)

        if self.note == "edit":
            self.initData()
    def initData(self):
        name = self.item.text(0)
        self.name.setText(name)
        if os.path.exists(self.path):
            with open(self.path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # 填入控件
            self.material_edit.setCurrentText(data.get("粉末材料", ""))
            self.fmd_edit.setText(data.get("粉末粒径(μm)", ""))
            self.sf_rate_edit.setValue(data.get("送粉转速(r/min)", 0))
            self.addition_rate_edit.setValue(data.get("粉末质量添加(g/min)", 0))
            self.carrier_gas_flow_edit.setText(data.get("粉末载气及流量(L/min)", ""))
            self.material_edit2.setCurrentText(data.get("送丝材料", ""))
            self.scd_edit.setValue(data.get("丝材直径(mm)", 0))
            self.ss_rate_edit.setValue(data.get("送丝速度(m/min)", 0))
            self.addition_rate_edit.setValue(data.get("丝质量添加(g/min)", 0))

            self.base_material_edit.setCurrentText(data.get("基板材料", ""))
            self.bansize_edit.setText(data.get("基板尺寸(mm)", ""))
            self.laser_power_edit.setValue(data.get("激光功率(W)", 0))
            self.welding_speed_edit.setValue(data.get("熔覆速度(mm/s)", 0))
            self.offset_edit.setValue(data.get("道间偏移(mm)", 0))
            self.lift_height_edit.setValue(data.get("层间抬升(mm)", 0))
            self.protect_gas_flow_edit.setText(data.get("保护气及流量(L/min)", ""))
            self.keep_time_edit.setValue(data.get("保护气保持时间(s)", 0))
            self.qd_flow_edit.setText(data.get("气刀及流量(L/min)", ""))
            self.pre_time_edit.setValue(data.get("加工前保护气时长(s)", 0))

    def get_folder_names(self, path):
        try:
            # 列出目录中的所有内容
            dir_contents = os.listdir(path)

            # 过滤出文件夹
            folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

            return folders
        except FileNotFoundError:
            print(f"路径 {path} 不存在")
            return []
        except Exception as e:
            print(f"遍历路径 {path} 时发生错误: {e}")
            return []
    def save_to_json(self):
        # 名称检查
        if self.note == "edit":
            path = self.item.parent().data(0, Qt.UserRole)
        else:
            path = self.item.data(0, Qt.UserRole)
        names = self.get_folder_names(path)
        if not self.name.text():
            QMessageBox.warning(self, "警告", "请输入工艺名称")
            return

        if self.note == "edit":
            ori_name = self.item.text(0)
            if self.name.text() != ori_name:
                if self.name.text() in names:
                    QMessageBox.warning(self, "警告", "该工艺名已存在")
                    return
            if os.path.exists(self.path):
                os.remove(self.path)
            file_path = osp.join(self.item.parent().data(0, Qt.UserRole), f"{self.name.text()}.json")
            self.item.setText(0, self.name.text())
            self.item.setData(0, Qt.UserRole, file_path)
        else:
            if self.name.text() in names:
                QMessageBox.warning(self, "警告", "该工艺名已存在")
                return
            new_item = QTreeWidgetItem(self.item)
            file_path = osp.join(self.item.data(0, Qt.UserRole), f"{self.name.text()}.json")
            new_item.setText(0, self.name.text())
            new_item.setData(0, Qt.UserRole, file_path)
            new_item.setData(0, Qt.UserRole + 1, "具体工艺")
            # 设置图标
            icon_path = os.path.join('..', 'resource', 'icon', 'doc.png')
            new_item.setIcon(0, QIcon(icon_path))
        data = {
            "工艺名称": self.name.text(),
            "粉末材料": self.material_edit.currentText(),
            "粉末粒径(μm)": self.fmd_edit.text(),
            "送粉转速(r/min)": self.ss_rate_edit.value(),
            "粉末质量添加(g/min)": self.addition_rate_edit.value(),
            "粉末载气及流量(L/min)": self.carrier_gas_flow_edit.text(),
            "送丝材料": self.material_edit2.currentText(),
            "丝材直径(mm)": self.scd_edit.value(),
            "送丝速度(m/min)": self.ss_rate_edit.value(),
            "丝质量添加(g/min)": self.addition_rate_edit.value(),
            "基板材料": self.base_material_edit.currentText(),
            "基板尺寸(mm)": self.bansize_edit.text(),
            "激光功率(W)": self.laser_power_edit.value(),
            "熔覆速度(mm/s)": self.welding_speed_edit.value(),
            "道间偏移(mm)": self.offset_edit.value(),
            "层间抬升(mm)": self.lift_height_edit.value(),
            "保护气及流量(L/min)": self.protect_gas_flow_edit.text(),
            "保护气保持时间(s)": self.keep_time_edit.value(),
            "气刀及流量(L/min)": self.qd_flow_edit.text(),
            "加工前保护气时长(s)": self.pre_time_edit.value(),
        }
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        self.accept()

class ProgramDialog(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        self.setWindowTitle("新建程序")
        self.setFixedSize(600, 350)  # (w,h)
        self.info_dict = mainWindow.info_dict
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")

        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)

        self.item = item
        self.note = "create"
        style = item.text(0)
        if self.item.data(0, Qt.UserRole + 1) == "具体程序":
            self.note = "edit"
            self.setWindowTitle("编辑程序")
            style = item.parent().text(0)
            self.path = item.parent().data(0, Qt.UserRole)
        else:
            self.path = item.data(0, Qt.UserRole)
        # 创建布局
        layout = QVBoxLayout()
        layout.setSpacing(30)
        layout1 = QHBoxLayout()
        layout2 = QHBoxLayout()
        form_layout = QFormLayout()
        vlayout1 = QFormLayout()
        vlayout1.setSpacing(20)  # 设置表单布局中控件间隙
        vlayout2 = QFormLayout()
        vlayout2.setSpacing(20)  # 设置表单布局中控件间隙

        # 输入字段
        self.name = QLineEdit()
        vlayout1.addRow("名称", self.name)

        self.type_combo = QLineEdit()
        self.type_combo.setText(style)
        self.type_combo.setReadOnly(True)
        vlayout2.addRow("类型", self.type_combo)

        self.layer_count_edit = QSpinBox()
        self.layer_count_edit.setRange(0, 10000)  # 设置范围
        vlayout1.addRow("层数", self.layer_count_edit)

        # 使用 QSpinBox 只允许输入整数
        self.path_count_edit = QSpinBox()
        self.path_count_edit.setRange(0, 10000)  # 设置范围
        vlayout2.addRow("道数", self.path_count_edit)

        self.laser_power_edit = QSpinBox()
        self.laser_power_edit.setRange(0, 10000)  # 设置范围
        # 创建 QLabel 并设置单位
        unit_label = QLabel("(W)")
        unit_label.setFixedWidth(52)
        # 创建 QHBoxLayout 将 QSpinBox 和 QLabel 放在一起
        jg_layout = QHBoxLayout()
        jg_layout.addWidget(self.laser_power_edit)  # 添加 QSpinBox
        jg_layout.addWidget(unit_label)  # 添加单位标签
        # 将水平布局添加到表单布局
        vlayout1.addRow("激光功率", jg_layout)  # 修改这里的标签
        # vlayout2.addRow("激光功率(W)", self.laser_power_edit)

        # 使用 QDoubleSpinBox 只允许输入浮点数
        self.melting_speed_edit = QSpinBox()
        self.melting_speed_edit.setRange(0, 10000)  # 设置范围
        # self.melting_speed_edit.setDecimals(2)  # 设置小数位数
        # 创建 QLabel 并设置单位
        unit_label = QLabel("(mm/s)")
        unit_label.setFixedWidth(52)
        # 创建 QHBoxLayout 将 QSpinBox 和 QLabel 放在一起
        rf_layout = QHBoxLayout()
        rf_layout.addWidget(self.melting_speed_edit)  # 添加 QSpinBox
        rf_layout.addWidget(unit_label)  # 添加单位标签
        # 将水平布局添加到表单布局
        vlayout2.addRow("熔覆速度", rf_layout)  # 修改这里的标签
        # vlayout2.addRow("熔覆速度(mm/s)", self.melting_speed_edit)

        # 连接信号
        self.layer_count_edit.valueChanged.connect(self.update_name)
        self.path_count_edit.valueChanged.connect(self.update_name)
        self.laser_power_edit.valueChanged.connect(self.update_name)
        self.melting_speed_edit.valueChanged.connect(self.update_name)

        layout1.addLayout(vlayout1)
        layout1.addLayout(vlayout2)

        layout.addLayout(layout1)
        layout.addLayout(form_layout)

        # 创建 QLabel 并设置文本
        label = QLabel("程序文件")
        layout2.addWidget(label)  # 将标签添加到布局
        # 程序文件列表
        self.file_list = QListWidget()
        layout2.addWidget(self.file_list)

        # 添加和删除按钮
        button_layout = QVBoxLayout()
        self.add_button = QPushButton("添加")
        self.remove_button = QPushButton("删除")
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.remove_button)
        layout2.addLayout(button_layout)
        layout.addLayout(layout2)

        # 确定和取消按钮
        self.ok_button = QPushButton("确定")
        self.cancel_button = QPushButton("取消")
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)

        # 设置主布局
        self.setLayout(layout)

        # 连接信号
        self.add_button.clicked.connect(self.add_file)
        self.remove_button.clicked.connect(self.remove_file)
        self.ok_button.clicked.connect(self.save_to_json)
        self.cancel_button.clicked.connect(self.reject)
        if self.note == "edit":
            self.initData()
            self.name.setText(self.item.text(0))
    def initData(self):
        filepath = osp.join(self.item.data(0, Qt.UserRole), "程序描述.json")
        # 确保文件存在
        if not osp.exists(filepath):
            QMessageBox.warning(self, "警告", f"程序描述文件不存在")
            return
        try:
            with open(filepath, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)
            # 将数据赋值到相应的控件
            self.name.setText(data["名称"])
            self.type_combo.setText(data["类型"])
            self.path_count_edit.setValue(data["道数"])
            self.layer_count_edit.setValue(data["层数"])
            self.laser_power_edit.setValue(data["激光功率(W)"])
            self.melting_speed_edit.setValue(data["熔覆速度(mm/s)"])
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法读取文件: {e}")
        pro_dir = self.item.data(0, Qt.UserRole)
        # 列出目录下的所有文件
        for filename in os.listdir(pro_dir):
            if not filename.endswith(".json"):
                self.file_list.addItem(filename)  # 添加文件名到 QListWidget
    def update_name(self):
        # 获取当前值并拼接
        layer_count = self.layer_count_edit.value()  # 获取层数
        dao_count = self.path_count_edit.value()  # 获取道数
        laser_power = self.laser_power_edit.value()
        melting_speed = self.melting_speed_edit.value()
        self.name.setText(f"L{layer_count}P{dao_count}-{laser_power}-{melting_speed}")  # 拼接并设置到名称中
    def add_file(self):
        # 打开文件对话框选择文件
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "选择图片文件", "",
                                                "Image Files (*.src *.dat);;All Files (*)",
                                                options=options)
        if files:  # 如果用户选择了文件
            for file in files:
                self.file_list.addItem(file)  # 将文件路径添加到 QListWidget

    def remove_file(self):
        # 删除选中的文件
        selected_items = self.file_list.selectedItems()
        for item in selected_items:
            self.file_list.takeItem(self.file_list.row(item))

    def save_to_json(self):
        # 名称检查
        name = self.name.text()
        names = get_folder_names(self.path)
        if not self.name.text():
            QMessageBox.warning(self, "警告", "请输入程序名称")
            return
        # 收集输入数据
        data = {
            "名称": self.name.text(),
            "类型": self.type_combo.text(),
            "道数": self.path_count_edit.value(),
            "层数": self.layer_count_edit.value(),
            "激光功率(W)": self.laser_power_edit.value(),
            "熔覆速度(mm/s)": self.melting_speed_edit.value()
        }
        # 设置文件名和路径
        dir_path = os.path.join(self.path, self.name.text())  # 目录路径
        file_path = os.path.join(dir_path, '程序描述.json')  # 完整文件路径
        # 新建模式先创建文件
        if self.note == "edit":
            ori_name = self.item.text(0)
            if name != ori_name:
                if name in names:
                    QMessageBox.warning(self, "警告", "该程序名已存在")
                    return
                # 重命名
                os.rename(osp.join(self.path, ori_name), osp.join(self.path, name))
                self.item.setText(0, name)
                self.item.setData(0, Qt.UserRole, osp.join(self.path, name))
        else:
            if name in names:
                QMessageBox.warning(self, "警告", "该程序名已存在")
                return
            # 创建目录（如果不存在的话）
            os.makedirs(dir_path, exist_ok=True)
        # 将数据写入 JSON 文件
        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
        # 遍历 QListWidget 中的所有项
        for index in range(self.file_list.count()):
            item = self.file_list.item(index)
            file_path = item.text()  # 获取文件路径

            if os.path.isfile(file_path):  # 确保路径是一个有效的文件
                # 检查文件路径是否为绝对路径
                if not os.path.isabs(file_path):
                    pass
                else:
                    # 复制文件到目标目录
                    try:
                        shutil.copy(file_path, dir_path)  # 复制文件
                        print(f"Copied {file_path} to {dir_path}")
                    except Exception as e:
                        print(f"Error copying {file_path}: {e}")
        self.accept()  # 关闭对话框

class ModelDialog(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        self.setWindowTitle("新建模型")
        self.setFixedSize(600, 400)
        self.mainWindow = mainWindow
        self.info_dict = mainWindow.info_dict
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        self.item = item
        self.note = "create"
        gy_style = self.item.parent().text(0)
        mx_style = self.item.text(0)
        if self.item.data(0, Qt.UserRole + 1) == "具体模型":
            self.note = "edit"
            self.setWindowTitle("编辑模型")
            gy_style = self.item.parent().parent().text(0)
            mx_style = self.item.parent().text(0)
            self.path = self.item.parent().data(0, Qt.UserRole)
        else:
            self.path = self.item.data(0, Qt.UserRole)
        # 创建主布局
        layout = QVBoxLayout()
        layout.setSpacing(15)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)
        hlayout1 = QHBoxLayout()
        hlayout2 = QHBoxLayout()
        vlayout1 = QFormLayout()
        vlayout1.setSpacing(15)  # 设置表单布局中控件间隙
        vlayout2 = QFormLayout()
        vlayout2.setSpacing(15)  # 设置表单布局中控件间隙

        # 创建输入字段
        self.name_edit = QLineEdit()
        vlayout1.addRow("名称", self.name_edit)

        self.gy_combo = QLineEdit()
        self.gy_combo.setText(gy_style)
        self.gy_combo.setReadOnly(True)
        vlayout2.addRow("工艺类型", self.gy_combo)

        self.type_combo = QLineEdit()
        self.type_combo.setText(mx_style)
        self.type_combo.setReadOnly(True)
        vlayout1.addRow("模型类型", self.type_combo)

        self.material_edit = QComboBox()
        self.material_edit.addItems(self.info_dict["具体材料"])
        vlayout2.addRow("熔覆材料", self.material_edit)

        self.laser_efficiency_edit = QSpinBox()
        self.laser_efficiency_edit.setRange(0, 10000)  # 设置范围
        # vlayout1.addRow("激光功率(W)", self.laser_efficiency_edit)
        # 创建 QLabel 并设置单位
        unit_label = QLabel("(W)")
        unit_label.setFixedWidth(52)
        # 创建 QHBoxLayout 将 QSpinBox 和 QLabel 放在一起
        jg_layout = QHBoxLayout()
        jg_layout.addWidget(self.laser_efficiency_edit)  # 添加 QSpinBox
        jg_layout.addWidget(unit_label)  # 添加单位标签
        # 将水平布局添加到表单布局
        vlayout1.addRow("激光功率", jg_layout)  # 修改这里的标签

        self.base_material_combo = QComboBox()
        self.base_material_combo.addItems(self.info_dict["具体材料"])
        vlayout2.addRow("基板材料", self.base_material_combo)

        self.melting_speed_edit = QSpinBox()
        self.melting_speed_edit.setRange(0, 10000)  # 设置范围
        # self.melting_speed_edit.setDecimals(2)  # 设置小数位数
        # 创建 QLabel 并设置单位
        unit_label1 = QLabel("(mm/s)")
        unit_label1.setFixedWidth(52)
        # 创建 QHBoxLayout 将 QSpinBox 和 QLabel 放在一起
        rf_layout = QHBoxLayout()
        rf_layout.addWidget(self.melting_speed_edit)  # 添加 QSpinBox
        rf_layout.addWidget(unit_label1)  # 添加单位标签
        # 将水平布局添加到表单布局
        vlayout1.addRow("熔覆速度", rf_layout)

        self.melting_shape_edit = QLineEdit()
        vlayout2.addRow("熔覆形式", self.melting_shape_edit)
        hlayout1.addLayout(vlayout1)
        hlayout1.addLayout(vlayout2)

        self.initial_condition_edit = QLineEdit()
        form_layout.addRow("初始条件", self.initial_condition_edit)

        self.boundary_condition_edit = QLineEdit()
        form_layout.addRow("边界条件", self.boundary_condition_edit)

        layout.addLayout(hlayout1)
        layout.addLayout(form_layout)

        # 创建 QLabel 并设置文本
        hlayout2.addWidget(QLabel("程序文件"))  # 将标签添加到布局
        # 程序文件列表
        self.file_list = QListWidget()
        hlayout2.addWidget(self.file_list)

        # 添加和删除按钮
        button_layout = QVBoxLayout()
        self.add_button = QPushButton("添加")
        self.remove_button = QPushButton("删除")
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.remove_button)
        hlayout2.addLayout(button_layout)
        layout.addLayout(hlayout2)

        # 确定和取消按钮
        self.ok_button = QPushButton("确定")
        self.cancel_button = QPushButton("取消")
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)

        # 设置主布局
        self.setLayout(layout)

        # 连接信号
        self.add_button.clicked.connect(self.add_file)
        self.remove_button.clicked.connect(self.remove_file)
        self.ok_button.clicked.connect(self.save_to_json)  # 修改此处来保存为 JSON
        self.cancel_button.clicked.connect(self.reject)
        if self.note == "edit":
            self.name_edit.setText(self.item.text(0))
            self.initData()
    def add_file(self):
        # 打开文件对话框选择文件
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "选择文件", "",
                                                "All Files (*)",
                                                options=options)
        if files:  # 如果用户选择了文件
            for file in files:
                self.file_list.addItem(file)  # 将文件路径添加到 QListWidget

    def initData(self):
        filepath = osp.join(self.item.data(0, Qt.UserRole), "模型描述.json")
        # 确保文件存在
        if not osp.exists(filepath):
            QMessageBox.warning(self, "警告", f"模型描述文件不存在")
            return
        try:
            with open(filepath, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)
            # 将数据赋值到相应的控件
            self.name_edit.setText(data["名称"])
            self.gy_combo.setText(data["工艺类型"])
            self.type_combo.setText(data["模型类型"])
            self.material_edit.setCurrentText(data["熔覆材料"])
            self.base_material_combo.setCurrentText(data["基板材料"])
            self.laser_efficiency_edit.setValue(data["激光功率(W)"])
            self.melting_speed_edit.setValue(data["熔覆速度(mm/s)"])
            self.melting_shape_edit.setText(data["熔覆形式"])
            self.initial_condition_edit.setText(data["初始条件"])
            self.boundary_condition_edit.setText(data["边界条件"])
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法读取文件: {e}")
        pro_dir = self.item.data(0, Qt.UserRole)

        # 列出目录下的所有文件
        for filename in os.listdir(pro_dir):
            if not filename.endswith(".json"):
                self.file_list.addItem(filename)  # 添加文件名到 QListWidget
    def remove_file(self):
        selected_items = self.file_list.selectedItems()
        for item in selected_items:
            self.file_list.takeItem(self.file_list.row(item))

    def save_to_json(self):
        # 名称检查
        name = self.name_edit.text()
        names = get_folder_names(self.path)
        if not self.name_edit.text():
            QMessageBox.warning(self, "警告", "请输入程序名称")
            return
        # 提取输入数据
        data = {
            "名称": self.name_edit.text(),
            "工艺类型": self.gy_combo.text(),
            "模型类型": self.type_combo.text(),
            "熔覆材料": self.material_edit.currentText(),
            "基板材料": self.base_material_combo.currentText(),
            "激光功率(W)": self.laser_efficiency_edit.value(),  # 确保为整数
            "熔覆速度(mm/s)": self.melting_speed_edit.value(),
            "熔覆形式": self.melting_shape_edit.text(),
            "初始条件": self.initial_condition_edit.text(),
            "边界条件": self.boundary_condition_edit.text()
        }

        # 设置文件名和路径
        dir_path = os.path.join(self.path, self.name_edit.text())  # 目录路径
        file_path = os.path.join(dir_path, '模型描述.json')  # 完整文件路径
        # 新建模式先创建文件
        if self.note == "edit":
            ori_name = self.item.text(0)
            if name != ori_name:
                if name in names:
                    QMessageBox.warning(self, "警告", "该模型名已存在")
                    return
                # 重命名
                os.rename(osp.join(self.path, ori_name), osp.join(self.path, name))
                self.item.setText(0, name)
                self.item.setData(0, Qt.UserRole, osp.join(self.path, name))
        else:
            if name in names:
                QMessageBox.warning(self, "警告", "该模型名已存在")
                return
            # 创建目录（如果不存在的话）
            os.makedirs(dir_path, exist_ok=True)
        # 将数据写入 JSON 文件
        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
        # 遍历 QListWidget 中的所有项
        for index in range(self.file_list.count()):
            item = self.file_list.item(index)
            file_path = item.text()  # 获取文件路径

            if os.path.isfile(file_path):  # 确保路径是一个有效的文件
                # 检查文件路径是否为绝对路径
                if not os.path.isabs(file_path):
                    pass
                else:
                    # 复制文件到目标目录
                    try:
                        shutil.copy(file_path, dir_path)  # 复制文件
                        print(f"Copied {file_path} to {dir_path}")
                    except Exception as e:
                        print(f"Error copying {file_path}: {e}")
        self.accept()  # 关闭对话框

class FilterCX(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        self.mainWindow = mainWindow
        self.style_path = osp.join(self.mainWindow.dataroot, "程序库")
        self.styles = self.get_folder_names(self.style_path)
        self.path = item.data(0, Qt.UserRole)
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        # self.data = self.load_data_from_json(osp.join(self.path, "程序目录.json"))
        self.data = []
        styles = self.get_folder_names(self.path)
        for style in styles:
            path = osp.join(self.path, style)
            jt_cxs = self.get_folder_names(path)
            for jt_cx in jt_cxs:
                cx_path = osp.join(path, jt_cx, "程序描述.json")
                data = self.read_json_file(cx_path)
                if data:
                    self.data.append(data)

        # 创建 UI 组件
        self.table_cx = QTableWidget()
        self.table_cx.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_cx.customContextMenuRequested.connect(self.show_context_menu)
        # 类型
        self.filter_type = QComboBox()
        self.filter_type.addItem("全部")
        self.filter_type.addItems(self.styles)
        # 层数
        self.filter_layer_min = QSpinBox()
        self.filter_layer_min.setRange(0, 10000)
        self.filter_layer_max = QSpinBox()
        self.filter_layer_max.setRange(0, 10000)
        self.filter_layer_max.setValue(10000)
        # 道数
        self.filter_dao_min = QSpinBox()
        self.filter_dao_min.setRange(0, 10000)
        self.filter_dao_max = QSpinBox()
        self.filter_dao_max.setRange(0, 10000)
        self.filter_dao_max.setValue(10000)
        # 功率
        self.filter_power_min = QSpinBox()
        self.filter_power_min.setRange(0, 10000)
        self.filter_power_max = QSpinBox()
        self.filter_power_max.setRange(0, 10000)
        self.filter_power_max.setValue(10000)
        # 熔覆速度
        self.filter_rfsd_min = QSpinBox()
        self.filter_rfsd_min.setRange(0, 10000)
        self.filter_rfsd_max = QSpinBox()
        self.filter_rfsd_max.setRange(0, 10000)
        self.filter_rfsd_max.setValue(10000)

        filter_button = QPushButton("筛选")
        filter_button.clicked.connect(self.apply_filter)

        # 布局设置
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("类型:"))
        filter_layout.addWidget(self.filter_type)
        filter_layout.addStretch(1)

        filter_layout.addWidget(QLabel("层数:"))
        filter_layout.addWidget(self.filter_layer_min)
        filter_layout.addWidget(QLabel("到"))
        filter_layout.addWidget(self.filter_layer_max)
        filter_layout.addStretch(1)

        filter_layout.addWidget(QLabel("道数:"))
        filter_layout.addWidget(self.filter_dao_min)
        filter_layout.addWidget(QLabel("到"))
        filter_layout.addWidget(self.filter_dao_max)
        filter_layout.addStretch(1)

        filter_layout.addWidget(QLabel("激光功率:"))
        filter_layout.addWidget(self.filter_power_min)
        filter_layout.addWidget(QLabel("到"))
        filter_layout.addWidget(self.filter_power_max)
        filter_layout.addStretch(1)

        filter_layout.addWidget(QLabel("熔覆速度:"))
        filter_layout.addWidget(self.filter_rfsd_min)
        filter_layout.addWidget(QLabel("到"))
        filter_layout.addWidget(self.filter_rfsd_max)
        filter_layout.addStretch(1)

        filter_layout.addWidget(filter_button)

        # 确定和取消按钮
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("确认")
        self.cancel_button = QPushButton("取消")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        layout = QVBoxLayout()
        layout.addLayout(filter_layout)
        # 设置大小策略
        layout.addWidget(self.table_cx)
        layout.addLayout(button_layout)
        self.setLayout(layout)
        self.selected_text = None  # 用于存储选中的项文本
        # 初始化表格
        self.populate_table(self.data)
        # 设置窗口标题和大小
        self.setWindowTitle("程序筛选")
        self.resize(1200, 600)

    def accept(self):
        item = self.table_cx.currentItem()
        row = item.row()  # 获取当前项的行索引
        col1_item = self.table_cx.item(row, 0)  # 前第一列
        col1_text = col1_item.text() if col1_item else "null"  # 获取项目名称
        col2_item = self.table_cx.item(row, 1)  # 前第二列
        col2_text = col2_item.text() if col2_item else "null"  # 获取项目类型
        self.selected_text = osp.join(col2_text, col1_text)
        super().accept()

    def get_selected_text(self):
        return self.selected_text  # 返回选中的项目文本
    def get_folder_names(self, path):
        try:
            # 列出目录中的所有内容
            dir_contents = os.listdir(path)

            # 过滤出文件夹
            folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

            return folders
        except FileNotFoundError:
            print(f"路径 {path} 不存在")
            return []
        except Exception as e:
            print(f"遍历路径 {path} 时发生错误: {e}")
            return []
    def read_json_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"文件 {file_path} 不存在")
            return None
        except json.JSONDecodeError as e:
            print(f"解析 JSON 文件 {file_path} 时发生错误: {e}")
            return None
    def show_context_menu(self, position):
        context_menu = QMenu(self)

        action_open = QAction("打开", self)
        action_open.triggered.connect(self.open_item)
        context_menu.addAction(action_open)

        # action_delete = QAction("导入", self)
        # action_delete.triggered.connect(self.imp_item)
        # context_menu.addAction(action_delete)

        context_menu.exec_(self.table_cx.viewport().mapToGlobal(position))
    def open_item(self):
        item = self.table_cx.currentItem()
        if item:
            row = item.row()  # 获取当前项的行索引
            # 获取该行前两列的内容
            if row >= 0:  # 确保行索引有效
                col1_item = self.table_cx.item(row, 0)  # 前第一列
                col2_item = self.table_cx.item(row, 1)  # 前第二列

                col1_text = col1_item.text() if col1_item else "null"
                col2_text = col2_item.text() if col2_item else "null"
                # 拼接路径
                path = os.path.join(self.path, col2_text, col1_text)

                # 打开文件
                if os.path.exists(path):  # 检查路径是否存在
                    try:
                        if os.name == 'nt':  # Windows
                            os.startfile(path)
                        elif os.name == 'posix':  # MacOS / Linux
                            subprocess.run(['open', path])  # MacOS
                            # subprocess.run(['xdg-open', path])  # Linux
                    except Exception as e:
                        print(f"打开文件时出错: {e}")
                else:
                    print(f"路径不存在: {path}")

    def imp_item(self):
        item = self.table_cx.currentItem()
        if item:
            row = item.row()
            self.table_cx.removeRow(row)
            print(f"删除项: {item.text()}")
    def populate_table(self, data):
        self.table_cx.setRowCount(len(data))
        self.table_cx.setColumnCount(len(data[0]))
        self.table_cx.setHorizontalHeaderLabels(data[0].keys())

        for row_index, row_data in enumerate(data):
            for col_index, (key, value) in enumerate(row_data.items()):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)  # 设置文本居中对齐
                self.table_cx.setItem(row_index, col_index, item)
        self.table_cx.setColumnWidth(0, 250)
        self.table_cx.setColumnWidth(1, 200)
        self.table_cx.setColumnWidth(2, 200)
        self.table_cx.setColumnWidth(3, 200)
        self.table_cx.setColumnWidth(4, 200)
        self.table_cx.horizontalHeader().setStretchLastSection(True)  # 最后一列扩展以占满空间
    def apply_filter(self):

        type_filter = self.filter_type.currentText()
        layer_min = self.filter_layer_min.value()
        layer_max = self.filter_layer_max.value()

        dao_min = self.filter_dao_min.value()
        dao_max = self.filter_dao_max.value()

        power_min = self.filter_power_min.value()
        power_max = self.filter_power_max.value()

        rfsd_min = self.filter_rfsd_min.value()
        rfsd_max = self.filter_rfsd_max.value()

        filtered_data = []

        for item in self.data:
            type_matches = (type_filter == "全部" or item["类型"] == type_filter)
            dao_matches = (dao_min <= item["道数"] <= dao_max)
            layer_matches = (layer_min <= item["层数"] <= layer_max)
            power_matches = (power_min <= item["激光功率(W)"] <= power_max)
            rfsd_matches = (rfsd_min <= item["熔覆速度(mm/s)"] <= rfsd_max)

            if dao_matches and type_matches and layer_matches and power_matches and rfsd_matches:
                filtered_data.append(item)

        self.populate_table(filtered_data)

class FilterMX(QDialog):
    def __init__(self, mainWindow, item):
        super().__init__()
        self.path = item.data(0, Qt.UserRole)
        self.info_dict = mainWindow.info_dict
        self.mainWindow = mainWindow
        self.style_path = osp.join(self.mainWindow.dataroot, "模型库")
        self.styles = self.get_folder_names(self.style_path)
        mx_styles_set = set()
        for i in range(len(self.styles)):
            path = osp.join(self.style_path, self.styles[i])  # 使用 i 而不是 0
            mx_styles = self.get_folder_names(path)
            for j in range(len(mx_styles)):
                mx_styles_set.add(mx_styles[j])  # 将获取的文件夹名称添加到集合中
        self.mx_styles = list(mx_styles_set)  # 转换为列表（可选）
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        # 设置字体大小
        font = QFont()
        font.setPointSize(10)  # 设置字体大小为10
        self.setFont(font)
        # self.data = self.load_data_from_json(osp.join(self.path, "模型目录.json"))
        self.data = []
        gy_styles = self.get_folder_names(self.path)
        for gy_style in gy_styles:
            path = osp.join(self.path, gy_style)
            mx_styles = self.get_folder_names(path)
            for mx_style in mx_styles:
                mx_style_path = osp.join(path, mx_style)
                jt_mxs = self.get_folder_names(mx_style_path)
                for jt_mx in jt_mxs:
                    mx_path = osp.join(mx_style_path, jt_mx, "模型描述.json")
                    if osp.exists(mx_path):  # 先检查文件是否存在
                        data = self.read_json_file(mx_path)
                        if data:
                            self.data.append(data)
                    else:
                        print(f"文件 {mx_path} 不存在")

        self.ColumnCount = len(self.data[0])
        self.keys = self.data[0].keys()
        # 创建 UI 组件
        self.table_cx = QTableWidget()
        self.table_cx.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_cx.customContextMenuRequested.connect(self.show_context_menu)
        # 创建筛选器组件
        self.filter_type = QComboBox()
        self.filter_type.addItem("全部")
        self.filter_type.addItems(self.styles)

        self.filter_type2 = QComboBox()
        self.filter_type2.addItem("全部")
        self.filter_type2.addItems(self.mx_styles)

        self.filter_material = QComboBox()  # 熔覆材料
        self.filter_material.addItem("全部")
        self.filter_material.addItems(self.info_dict["具体材料"])
        self.filter_substrate = QComboBox()  # 基板材料
        self.filter_substrate.addItem("全部")
        self.filter_substrate.addItems(self.info_dict["具体材料"])
        # 激光功率
        self.filter_jg_min = QSpinBox()
        self.filter_jg_min.setRange(0, 10000)
        self.filter_jg_max = QSpinBox()
        self.filter_jg_max.setRange(0, 10000)
        self.filter_jg_max.setValue(10000)  # 设置默认值为最大值
        # 熔覆速度
        self.filter_speed_min = QSpinBox()
        self.filter_speed_min.setRange(0, 10000)
        self.filter_speed_max = QSpinBox()
        self.filter_speed_max.setRange(0, 10000)
        self.filter_speed_max.setValue(10000)  # 设置默认值为最大值

        filter_button = QPushButton("筛选")
        filter_button.clicked.connect(self.apply_filter)

        # 布局设置
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("工艺类型:"))
        filter_layout.addWidget(self.filter_type)
        filter_layout.addStretch(1)
        filter_layout.addWidget(QLabel("模型类型:"))
        filter_layout.addWidget(self.filter_type2)
        filter_layout.addStretch(1)
        filter_layout.addWidget(QLabel("熔覆材料:"))
        filter_layout.addWidget(self.filter_material)
        filter_layout.addStretch(1)
        filter_layout.addWidget(QLabel("基板材料:"))
        filter_layout.addWidget(self.filter_substrate)
        filter_layout.addStretch(1)
        filter_layout.addWidget(QLabel("激光功率:"))
        filter_layout.addWidget(self.filter_jg_min)
        filter_layout.addWidget(QLabel("到"))
        filter_layout.addWidget(self.filter_jg_max)
        filter_layout.addStretch(1)
        filter_layout.addWidget(QLabel("熔覆速度:"))
        filter_layout.addWidget(self.filter_speed_min)
        filter_layout.addWidget(QLabel("到"))
        filter_layout.addWidget(self.filter_speed_max)
        filter_layout.addWidget(filter_button)
        # 确定和取消按钮
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("确认")
        self.cancel_button = QPushButton("取消")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        # 主布局
        main_layout = QVBoxLayout()
        main_layout.addLayout(filter_layout)
        # 设置大小策略
        self.table_cx.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(self.table_cx)
        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)

        # 初始化表格
        self.populate_table(self.data)

        # 设置窗口标题和大小
        self.setWindowTitle("模型筛选")
        self.resize(800, 600)

    def accept(self):
        item = self.table_cx.currentItem()
        row = item.row()  # 获取当前项的行索引
        col0_item = self.table_cx.item(row, 0)
        col0_text = col0_item.text() if col0_item else "null"  # 获取工艺类型
        col1_item = self.table_cx.item(row, 1)  # 前第二列
        col1_text = col1_item.text() if col1_item else "null"  # 获取工艺类型
        col2_item = self.table_cx.item(row, 2)  # 前第三列
        col2_text = col2_item.text() if col2_item else "null"  # 获取模型类型
        self.selected_text = [col0_text, col1_text, col2_text]
        super().accept()

    def get_selected_text(self):
        return self.selected_text  # 返回选中的项目文本
    def get_folder_names(self, path):
        try:
            # 列出目录中的所有内容
            dir_contents = os.listdir(path)

            # 过滤出文件夹
            folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

            return folders
        except FileNotFoundError:
            print(f"路径 {path} 不存在")
            return []
        except Exception as e:
            print(f"遍历路径 {path} 时发生错误: {e}")
            return []

    def read_json_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"文件 {file_path} 不存在")
            return None
        except json.JSONDecodeError as e:
            print(f"解析 JSON 文件 {file_path} 时发生错误: {e}")
            return None
    def populate_table(self, data):
        self.table_cx.setRowCount(len(data))
        self.table_cx.setColumnCount(self.ColumnCount)
        self.table_cx.setHorizontalHeaderLabels(self.keys)
        for row_index, row_data in enumerate(data):
            for col_index, (key, value) in enumerate(row_data.items()):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)  # 设置文本居中对齐
                self.table_cx.setItem(row_index, col_index, item)
        self.table_cx.setColumnWidth(0, 220)
        self.table_cx.horizontalHeader().setStretchLastSection(True)  # 最后一列扩展以占满空间

    def show_context_menu(self, position):
        context_menu = QMenu(self)

        action_open = QAction("打开", self)
        action_open.triggered.connect(self.open_item)
        context_menu.addAction(action_open)

        # action_delete = QAction("导入", self)
        # action_delete.triggered.connect(self.imp_item)
        # context_menu.addAction(action_delete)

        context_menu.exec_(self.table_cx.viewport().mapToGlobal(position))

    def open_item(self):
        item = self.table_cx.currentItem()
        if item:
            row = item.row()  # 获取当前项的行索引
            # 获取该行前两列的内容
            if row >= 0:  # 确保行索引有效
                col1_item = self.table_cx.item(row, 0)  # 第一列
                col2_item = self.table_cx.item(row, 1)  # 第二列
                col3_item = self.table_cx.item(row, 2)  # 第三列
                col1_text = col1_item.text()
                col2_text = col2_item.text()
                col3_text = col3_item.text()
                # 拼接路径
                path = os.path.join(self.path, col2_text, col3_text, col1_text)

                # 打开文件
                if os.path.exists(path):  # 检查路径是否存在
                    try:
                        if os.name == 'nt':  # Windows
                            os.startfile(path)
                        elif os.name == 'posix':  # MacOS / Linux
                            subprocess.run(['open', path])  # MacOS
                            # subprocess.run(['xdg-open', path])  # Linux
                    except Exception as e:
                        print(f"打开文件时出错: {e}")
                else:
                    print(f"路径不存在: {path}")

    def imp_item(self):
        item = self.table_cx.currentItem()
        if item:
            row = item.row()
            self.table_cx.removeRow(row)
            print(f"删除项: {item.text()}")

    def apply_filter(self):
        type_filter = self.filter_type.currentText()
        type2_filter = self.filter_type2.currentText()
        material_filter = self.filter_material.currentText()
        substrate_filter = self.filter_substrate.currentText()
        jg_min = self.filter_jg_min.value()
        jg_max = self.filter_jg_max.value()
        speed_min = self.filter_speed_min.value()
        speed_max = self.filter_speed_max.value()

        filtered_data = []

        for item in self.data:
            type_matches = (type_filter == "全部" or item["工艺类型"] == type_filter)
            type2_matches = (type2_filter == "全部" or item["模型类型"] == type2_filter)
            material_matches = (material_filter == "全部" or item["熔覆材料"] == material_filter)
            substrate_matches = (substrate_filter == "全部" or item["基板材料"] == substrate_filter)
            jg_matches = (jg_min <= item["激光功率(W)"] <= jg_max)
            speed_matches = (speed_min <= item["熔覆速度(mm/s)"] <= speed_max)

            if type_matches and type2_matches and material_matches and substrate_matches and jg_matches and speed_matches:
                filtered_data.append(item)

        self.populate_table(filtered_data)

class SaveAsDialog(QDialog):
    def __init__(self, mainWindows):
        super().__init__()
        # 设置窗口图标
        icon_path = os.path.join('..', 'resource', 'icon', 'PyDracula.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            print(f"Icon file not found: {icon_path}")
        self.root = mainWindows.root
        self.datapath = osp.join(self.root, '数据库')
        self.projectpath = osp.join(self.root, '项目库')
        self.pro_list = self.get_folder_names(self.projectpath)
        self.savepath = './'
        self.setWindowTitle("数据库另存")
        self.setFixedSize(400, 300)  # 设置窗口大小
        self.setStyleSheet("font-size: 11pt; background-color: rgb(240, 240, 240);")  # 设置字体大小和背景色
        self.initUI()

    def initUI(self):
        # 主布局
        layout = QHBoxLayout()

        # 创建复选框区
        left_layout = QVBoxLayout()
        self.checkboxes = [
            QCheckBox("材料库"),
            QCheckBox("设备库"),
            QCheckBox("工艺库"),
            QCheckBox("程序库"),
            QCheckBox("模型库"),
        ]

        # 添加复选框到布局
        for checkbox in self.checkboxes:
            left_layout.addWidget(checkbox)

        # 创建保存路径输入框和按钮
        self.path_input = QLineEdit()
        browse_button = QPushButton("浏览")
        browse_button.clicked.connect(self.browse)

        path_layout = QHBoxLayout()
        path_layout.addWidget(QLabel("保存路径"))
        path_layout.addWidget(self.path_input)
        path_layout.addWidget(browse_button)

        layout.addLayout(left_layout)

        # 添加比例控制的SpacerItem
        layout.addItem(QSpacerItem(70, 0, QSizePolicy.Fixed, QSizePolicy.Expanding))
        # layout.addLayout(right_layout)
        # layout.addLayout(left_layout)

        # 创建项目库列表框
        self.list_widget = QListWidget()
        self.list_widget.addItems(self.pro_list)
        self.list_widget.setSelectionMode(QListWidget.MultiSelection)

        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("项目库"))
        right_layout.addWidget(self.list_widget)

        # 创建确认和取消按钮
        button_layout = QHBoxLayout()
        confirm_button = QPushButton("确定")
        cancel_button = QPushButton("取消")
        button_layout.addWidget(confirm_button)
        button_layout.addWidget(cancel_button)

        layout.addLayout(right_layout)
        all_layout = QVBoxLayout()
        all_layout.addLayout(layout)
        all_layout.addLayout(path_layout)
        all_layout.addLayout(button_layout)
        self.setLayout(all_layout)

        # 连接按钮事件
        confirm_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)

    def get_folder_names(self, path):
        try:
            # 列出目录中的所有内容
            dir_contents = os.listdir(path)

            # 过滤出文件夹
            folders = [name for name in dir_contents if osp.isdir(osp.join(path, name))]

            return folders
        except FileNotFoundError:
            print(f"路径 {path} 不存在")
            return []
        except Exception as e:
            print(f"遍历路径 {path} 时发生错误: {e}")
            return []
    def accept(self):
        # 获取保存路径
        file_path = self.path_input.text()
        if not file_path:
            QMessageBox.warning(self, "警告", "请先选择保存路径！")
            return
        file_path = osp.join(file_path, 'database')
        data_path = osp.join(file_path, '数据库')
        pro_path = osp.join(file_path, '项目库')
        libraries = ['材料库', '设备库', '工艺库', '程序库', '模型库']
        # os.makedirs(file_path, exist_ok=True)  # 创建数据库文件夹，若已存在则不会抛出异常
        os.makedirs(data_path, exist_ok=True)
        os.makedirs(pro_path, exist_ok=True)
        # 创建子文件夹
        for library in libraries:
            library_folder_path = os.path.join(data_path, library)
            os.makedirs(library_folder_path, exist_ok=True)
        selected_libraries = []
        for checkbox in self.checkboxes:
            if checkbox.isChecked():
                selected_libraries.append(checkbox.text())

        if not selected_libraries:
            QMessageBox.warning(self, "警告", "请至少选择一个库！")
            return

        # 保存选中的库到文件夹
        for library in selected_libraries:
            path1 = osp.join(self.datapath, library)
            path2 = osp.join(data_path, library)
            print(path2)
            # 如果目标文件夹已存在，先删除
            if os.path.exists(path2):
                shutil.rmtree(path2)
            # 复制文件夹
            shutil.copytree(path1, path2)

        selected_items = self.list_widget.selectedItems()  # 获取所有选中的项目
        selected_names = [item.text() for item in selected_items]  # 提取每个项目的名称
        # 保存选中的库到文件夹
        for library in selected_names:
            path1 = osp.join(self.projectpath, library)
            path2 = osp.join(pro_path, library)
            # 如果目标文件夹已存在，先删除
            if os.path.exists(path2):
                shutil.rmtree(path2)
            # 复制文件夹
            shutil.copytree(path1, path2)

        msg_box = QMessageBox.information(self, "成功", "选中的库和项目已经保存！")
        if msg_box == QMessageBox.Ok:
            super().accept()  # 关闭对话框

    def browse(self):
        # 打开文件夹选择对话框
        options = QFileDialog.Options()
        folder_name = QFileDialog.getExistingDirectory(self, "选择保存路径", "", options=options)
        if folder_name:
            self.path_input.setText(folder_name)


class AddPropertyDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("添加材料属性")  # 设置标题

        # 创建布局
        layout = QFormLayout(self)

        # 添加属性名称输入框
        self.property_name = QLineEdit()
        layout.addRow(QLabel("属性名"), self.property_name)

        # 第一列
        self.first_column = QLineEdit()
        layout.addRow(QLabel("第一列标题"), self.first_column)
        # 第二列
        self.second_column = QLineEdit()
        layout.addRow(QLabel("第二列标题"), self.second_column)
        # 第三列
        self.third_column = QLineEdit()
        layout.addRow(QLabel("第三列标题"), self.third_column)

        # 确定和取消按钮
        button_layout = QHBoxLayout()
        self.confirm_button = QPushButton("确定")
        self.cancel_button = QPushButton("取消")

        button_layout.addWidget(self.confirm_button)
        button_layout.addWidget(self.cancel_button)

        layout.addRow(button_layout)

        # 连接按钮事件
        self.cancel_button.clicked.connect(self.reject)
        self.confirm_button.clicked.connect(self.accept)
    def accept(self):
        self.list_data = [self.property_name.text(), self.first_column.text(), self.second_column.text(), self.third_column.text()]
        super().accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    dialog = FilterMX('55')
    if dialog.exec_() == QDialog.Accepted:
        # 处理确认后的逻辑
        print("对话框确认")
    sys.exit(app.exec_())
